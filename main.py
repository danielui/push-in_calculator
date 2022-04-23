from openpyxl import load_workbook


def calculate_ei(d, d1, d2, pmin, Rz1, Rz2, E1, E2, v1, v2, ES):
    sg1 = (d**2 + d1**2) / (d**2 - d1**2)
    sg2 = (d2**2 + d**2) / (d2**2 - d**2)
    Wmin = ((sg2 + v2) / E2 +
            (sg1 - v1) / E1) * d * pmin
    WminRZ = Wmin + 2 * 0.7 * (Rz1 + Rz2)
    ei = WminRZ + ES
    # return ei in um
    return ei * 10**6


def find_shaft_tolerance(ei, eiRow):
    wb = load_workbook(filename=r"tolerances.xlsx")

    for i in range(2, 6):
        sheet = wb[str(i)]
        for row in sheet.iter_rows(min_row=eiRow, max_row=eiRow, min_col=4, max_col=14):
            for cell in row:
                if isinstance(cell.value, int):
                    if int(cell.value) >= ei:
                        return sheet[2][(cell.column) - 1].value

    return "Sorry, there is no normalized tolerance for your ei"


def find_ES(d, holeTolerance):
    wb = load_workbook(
        filename=r"tolerances.xlsx")
    sheet = wb["1"]
    switch = {'H6': 'D', 'H7': 'E', 'H8': 'F', 'H9': 'G',
              'H10': 'H', 'H11': 'I', 'H12': 'J', 'H13': 'K', 'H14': 'L'}

    for i in range(5, 55):
        if d <= int(sheet["A" + str(i)].value):
            row = i - 2
            # return ES in um, return row number for further operations
            return (sheet[str(switch[str(holeTolerance)]) + str(row)].value), i - 1


def calculate_load(d, ln, mi, loadType, P, Ms, Mg):

    if loadType == "axial":
        pmin = (1.7 * P) / (mi * 3.14 * ln * d)
        return pmin

    elif loadType == "twisting_moment":
        pmin = (3.4 * Ms) / (mi * 3.14 * ln * (d**2))
        return pmin

    elif loadType == "axial_and_twisting_moment":
        pmin = ((P ** 2 + (2 * Ms / d) ** 2) ** 0.5) / (mi * 3.14 * ln * d)
        return pmin

    elif loadType == "bending_moment":
        pmin = (3 * Mg) / ((ln**2) * (d**2) * 0.7)
        return pmin


def get_material_data(userMaterial):
    with open(r"materials.txt") as materials:
        materialListLength = len(materials.readlines())
    with open(r"materials.txt") as materials:
        for _ in range(materialListLength):
            materialData = materials.readline().split()
            if materialData[0] == str(userMaterial):
                return materialData
        # if no material in file, post new material
        print("No such material in the file \n Please add it")
        return post_material_data(input("new material name: "), input("new material Young's module: "),
                                  input("new material Poisson ratio: "))


def post_material_data(newMaterialName, newMaterialYoungModule, newMaterialPoissonRatio):
    with open(r"materials.txt") as materials:
        materialList = (materials.readlines())
        materialCheckList = [i.split(" ") for i in materialList]
        flatMaterialCheckList = [
            item for sublist in materialCheckList for item in sublist]
        if newMaterialName not in flatMaterialCheckList:
            with open(r"materials.txt", mode="a") as materials:
                materials.write('\n' + str(newMaterialName) + " " +
                                str(newMaterialYoungModule) + " " + str(newMaterialPoissonRatio))
        return get_material_data(newMaterialName)


def main():

    d = float(input("Push-in connection diameter[1-500][mm]: "))
    while (d < 1) or (d > 500):
        print("Diameter out of range, please type again")
        d = float(input("Type push-in connection diameter[1-500][mm]: "))
    d = d / 1000  # converts mm into m
    d1 = float(input("Shaft axial hole diameter[mm]: "))
    while (d1 < 0) or (d1 >= (d * 1000)):
        print("Diameter out of range, please type again")
        d1 = float(input("Shaft axial hole diameter[mm]: "))
    d1 = d1 / 1000  # converts mm into m
    d2 = float(input("Hub external diameter[mm]: "))
    d2 = d2 / 1000  # converts mm into m
    ln = float(input("Connection length[mm]: "))
    ln = ln / 1000  # converts mm into m
    Rz1 = 0.4
    Rz1 = Rz1 / (10**6)  # converts um into m
    Rz2 = 0.4
    Rz2 = Rz2 / (10**6)  # converts um into m
    shaftMaterial = get_material_data(input("Shaft material: "))
    hubMaterial = get_material_data(input("Hub material: "))
    ES = find_ES(
        d * 1000, holeTolerance=(input("Hub hole tolerance:")).upper())
    mi = float(input("Friction coefficient between " +
                     str(shaftMaterial[0] + " " "and" + " " + str(hubMaterial[0] + " "))))

    loadDict = {'1': 'axial', '2': 'twisting_moment', '3': 'axial_and_twisting_moment',
                '4': 'bending_moment'}
    loadChoice = input(
        'Choose load type: \n 1 - axial\n 2 - twisting moment\n 3 - axial and twisting moment\n 4 - bending moment  ')
    if loadChoice == "1":
        P = float(input("Force [N]: "))
        Ms = 0
        Mg = 0
    elif loadChoice == "2":
        P = 0
        Ms = float(input("Torque [N*m]: "))
        Mg = 0
    elif loadChoice == "3":
        P = float(input("Force [N*m]: "))
        Ms = float(input("Torque [N*m]: "))
        Mg = 0
    elif loadChoice == "4":
        P = 0
        Ms = 0
        Mg = float(input("Torque [N*m]: "))

    pmin = calculate_load(d, ln, mi, loadDict[loadChoice], P, Ms, Mg)
    ei = calculate_ei(d, d1, d2, pmin, Rz1, Rz2,
                      float(shaftMaterial[1]), float(hubMaterial[1]),
                      float(shaftMaterial[2]), float(hubMaterial[2]),
                      (ES[0])/(10**6))
    shaftTolerance = find_shaft_tolerance(ei, ES[1])

    print("shaft tolerance = " + str(shaftTolerance))
    print("ei[um] = " + str(ei))


main()
